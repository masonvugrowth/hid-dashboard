"""Government visitor data router — CRUD + Excel import."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.gov_visitor import GovVisitorData
from app.routers.auth import require_admin

router = APIRouter()

MONTHS = ["jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec"]


def _envelope(data):
    return {"success": True, "data": data, "error": None,
            "timestamp": datetime.now(timezone.utc).isoformat()}


class GovVisitorIn(BaseModel):
    destination: str
    source_country: str
    rank: Optional[int] = None
    jan: int = 0
    feb: int = 0
    mar: int = 0
    apr: int = 0
    may: int = 0
    jun: int = 0
    jul: int = 0
    aug: int = 0
    sep: int = 0
    oct: int = 0
    nov: int = 0
    dec: int = 0
    total: int = 0
    data_year: Optional[int] = None


def _row_to_dict(r: GovVisitorData) -> dict:
    return {
        "id": str(r.id),
        "destination": r.destination,
        "source_country": r.source_country,
        "rank": r.rank,
        **{m: getattr(r, m) or 0 for m in MONTHS},
        "total": r.total or 0,
        "data_year": r.data_year,
    }


@router.get("")
def list_gov_visitor(
    destination: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(GovVisitorData)
    if destination:
        q = q.filter(GovVisitorData.destination == destination)
    q = q.order_by(GovVisitorData.destination, GovVisitorData.rank.asc().nullslast())
    rows = q.all()
    return _envelope([_row_to_dict(r) for r in rows])


@router.get("/destinations")
def list_destinations(db: Session = Depends(get_db)):
    rows = db.query(GovVisitorData.destination).distinct().order_by(GovVisitorData.destination).all()
    return _envelope([r[0] for r in rows])


@router.post("")
def create_gov_visitor(
    body: GovVisitorIn,
    _admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    row = GovVisitorData(**body.model_dump())
    db.add(row)
    db.commit()
    db.refresh(row)
    return _envelope(_row_to_dict(row))


@router.delete("/{row_id}")
def delete_gov_visitor(
    row_id: str,
    _admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    row = db.query(GovVisitorData).filter_by(id=row_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Row not found")
    db.delete(row)
    db.commit()
    return _envelope({"deleted": True})


@router.delete("")
def delete_by_destination(
    destination: str = Query(...),
    _admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    count = db.query(GovVisitorData).filter_by(destination=destination).delete()
    db.commit()
    return _envelope({"deleted_count": count})


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    _admin=Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files accepted")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    imported = 0
    destinations_imported = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        destination = sheet_name.strip()

        # Delete existing data for this destination before re-importing
        db.query(GovVisitorData).filter_by(destination=destination).delete()

        rows_list = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows_list) < 2:
            continue

        # Header row: [Unnamed/rank, destination_name, Jan, Feb, ..., Dec, Sum]
        for row in rows_list[1:]:
            if not row or len(row) < 15:
                continue
            rank_val = row[0]
            country = row[1]
            if not country or str(country).strip() == "":
                continue

            month_vals = []
            for i in range(2, 14):
                v = row[i]
                month_vals.append(int(v) if v and str(v).strip() not in ("", "-") else 0)

            total_val = row[14] if len(row) > 14 else sum(month_vals)
            total_val = int(total_val) if total_val and str(total_val).strip() not in ("", "-") else sum(month_vals)

            record = GovVisitorData(
                destination=destination,
                source_country=str(country).strip(),
                rank=int(rank_val) if rank_val else None,
                jan=month_vals[0], feb=month_vals[1], mar=month_vals[2],
                apr=month_vals[3], may=month_vals[4], jun=month_vals[5],
                jul=month_vals[6], aug=month_vals[7], sep=month_vals[8],
                oct=month_vals[9], nov=month_vals[10], dec=month_vals[11],
                total=total_val,
            )
            db.add(record)
            imported += 1

        destinations_imported.append(destination)

    db.commit()
    wb.close()

    return _envelope({
        "imported_rows": imported,
        "destinations": destinations_imported,
        "filename": file.filename,
    })
